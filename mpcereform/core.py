"""Functions for copying data from the old database to the new."""

#pylint:disable=line-too-long;
#pylint:disable=too-many-lines;

import re
from importlib.resources import read_text, path
from uuid import uuid1

import mysql.connector as mysql
from openpyxl import load_workbook

from mpcereform.utils import parse_date

class LocalDB():
    """Class for managing connection to database server"""

    UNCHANGED_TABLES = {
        # Tables identical to their STN counterparts
        'mpce.stn_client_profession': 'manuscripts.clients_professions',
        'mpce.stn_edition_call_number': 'manuscripts.books_call_numbers',
        'mpce.stn_edition_catalogue': 'manuscripts.books_stn_catalogues',
        'mpce.stn_client_correspondence_ms': 'manuscripts.clients_correspondence_manuscripts',
        'mpce.stn_client_correspondence_place': 'manuscripts.clients_correspondence_places',
        'mpce.stn_order': 'manuscripts.orders',
        'mpce.stn_order_agent': 'manuscripts.orders_agents',
        'mpce.stn_order_sent_via': 'manuscripts.orders_sent_via',
        'mpce.stn_order_sent_via_place': 'manuscripts.orders_sent_via_place',
        'mpce.stn_transaction_volumes_exchanged': 'manuscripts.transactions_volumes_exchanged'
    }

    TRANSACTION_CODING = {
        # Codes for STN transactions
        'in':1,
        'in (printing)':1,
        'out':2,
        'out (free gifts)':2,
        'out (transfer)':2,
        'in (transfer)':1,
        'in (return)':1,
        'out (lost)':2,
        'out (return)':2,
        'stock take':3,
        'out (other)':2,
        'in (found)':1,
        'out (profit and loss)':2,
        'in (recount)':1,
        'out (missing)':2,
        'in (other)':1,
        'out (binders)':2,
        'bilan de sortie':3,
        'bilan':3,
        'out (faults)':2,
        'in (profit and loss)':1,
        'out (inferred)':2,
        'sales ms1003':3,
        'returned before arrival':1,
        'in (rétiré)':1,
        'catalogue':3,
        'out (durand commission)':2,
        'commission':3,
        'in (durand commission)':1,
        'in (assumed printing)':1,
        'bilan (adjustment)':3
    }

    DARNTON_CLIENTS = {
        # Key for client information in Robert Darnton's STN sample
        'Bergeret': 'cl0335',
        'Buchet': 'cl1468',
        'Caldesaigues': 'cl1203',
        'Charmet': 'cl0274',
        'Chevrier': 'cl1763',
        'Couret de Villeneuve': 'cl1510',
        'Fontanel': 'cl1224',
        'Gaude': 'cl1472',
        'Lair': 'cl0328',
        'Lepagnez': 'cl0288',
        'Letourmy': 'cl1514',
        'Mossy': 'cl1210',
        'Pavie': 'cl1834',
        'Rigaud': 'cl1266',
        'Robert&Gauthier': 'cl0353',
        'Sens': 'cl2067'
    }

    def __init__(self, user='root', host='127.0.0.1', password=None):
        self.conn = mysql.connect(user=user, host=host, password=password)

        # Check databases exist
        cur = self.conn.cursor()
        cur.execute("SHOW DATABASES")
        db_list = [x[0] for x in cur.fetchall()]

        def check_for_dbs(msg=None):
            # Check for mpce
            if 'mpce' in db_list:
                if msg is None:
                    msg = "Existing MPCE database found. Overwrite? [y/n] "
                resp = input(msg)
                if resp == 'y':
                    print("Overwriting existing database...")
                    cur.execute("DROP DATABASE mpce")
                    self.create_new_db()
                elif resp == 'n':
                    pass
                else:
                    check_for_dbs("Overwrite existing MPCE database? Type 'y' or 'n': ")
            else:
                print("MPCE database not found. Creating it...")
                self.create_new_db()

            # Check for manuscripts
            if 'manuscripts' not in db_list:
                raise mysql.DatabaseError("Manuscripts database not found!")

        check_for_dbs()

    def create_new_db(self):
        """Rebuilds the new MPCE database from schema"""

        # Read in schema
        schema_raw = read_text('mpcereform.sql', 'mpce_database.sql')

        # Strip multiline comments
        # Use a non-greedy match, so it will find each seperate comment
        sql_com_rgx = re.compile(r'/\*.+?\*/', re.DOTALL)
        schema_stripped = sql_com_rgx.sub('', schema_raw)

        # Split on semicolons
        schema_split = schema_stripped.split(';')

        # Execute
        cur = self.conn.cursor()
        for stmt in schema_split:
            cur.execute(stmt)
        self.conn.commit()
        cur.close()

    def import_works(self):
        """Copies works from old db to new"""

        # New cursor
        cur = self.conn.cursor()

        # Copy basic data directly from manuscript_books
        print("Importing works...")
        cur.execute("""
            INSERT INTO mpce.work (
                work_code, work_title, parisian_keyword, illegality_notes
            )
            SELECT super_book_code, super_book_title, parisian_keyword, illegality
            FROM manuscripts.manuscript_books
        """)
        self.conn.commit()
        print(f'{cur.rowcount} works copied.')

        # Copy categorisation data
        print("Processing keywords...")
        cur.execute("""
            UPDATE mpce.work AS w, manuscripts.manuscript_cat_fuzzy AS cf
            SET
                w.categorisation_fuzzy_value = cf.fuzzyValue,
                w.categorisation_notes = cf.fuzzyComment
            WHERE w.work_code = cf.super_book_code AND
            cf.fuzzyValue <> 'null'
        """)

        # Reform the 20 or so invalid keyword codes
        cur.execute('SELECT * FROM manuscripts.keywords')
        all_keywords = cur.fetchall()
        next_keyid = max([int(code[1:])
                          for (code, word, definition, tag) in all_keywords if len(code) == 5])
        # Create map
        cur.execute("""
            CREATE TEMPORARY TABLE mpce.keyword_map (
                old_code VARCHAR(255), new_code VARCHAR(255)
            )""")
        keyword_map = []
        for code, _, _, _ in all_keywords:
            if len(code) != 5:
                next_keyid += 1
                keyword_map.append((code, f'k{next_keyid}'))
            else:
                keyword_map.append((code, code))
        cur.executemany("""
            INSERT INTO mpce.keyword_map
            VALUES (%s, %s)
        """, seq_params=keyword_map)
        cur.execute("""
            INSERT INTO mpce.keyword
            SELECT map.new_code, kw.keyword, kw.definition, kw.tag_code
            FROM manuscripts.keywords AS kw
                LEFT JOIN mpce.keyword_map AS map
                    ON kw.keyword_code = map.old_code
        """)

        # Break keywords out into join table
        # Keyword assignments are comma-seperated values in 'manuscripts'
        cur.execute("""
            SELECT super_book_code, keywords
            FROM manuscripts.manuscript_books
            WHERE CHAR_LENGTH(keywords) > 1
        """)
        keywords = cur.fetchall()
        keywords_split = []
        for sbk, kwds in keywords:
            for kwd in kwds.split(','):
                keywords_split.append((sbk, kwd.strip()))
        cur.execute("""CREATE TEMPORARY TABLE mpce.work_keyword_temp (
            work_code VARCHAR(255),
            keyword_code VARCHAR(255)
        )
        """)
        cur.executemany("""
            INSERT INTO mpce.work_keyword_temp (work_code, keyword_code)
            VALUES (%s, %s)
        """, seq_params=keywords_split)
        cur.execute("""
            INSERT INTO mpce.work_keyword (work_code, keyword_code)
            SELECT temp.work_code, map.new_code
            FROM mpce.work_keyword_temp AS temp
                LEFT JOIN mpce.keyword_map AS map
                    ON temp.keyword_code = map.old_code
        """)
        self.conn.commit()
        print(f'{cur.rowcount} keyword assignments copied.')

        # Import rest of keyword data
        cur.execute("""
            INSERT INTO mpce.parisian_category
            SELECT * FROM manuscripts.parisian_keywords
        """)
        cur.execute("""
            INSERT INTO mpce.tag
            SELECT * FROM manuscripts.tags
        """)
        self.conn.commit()
        print('Parisian categories, keywords and tags imported.')

        # Import keyword associations (need some massaging)
        cur.execute("""
            INSERT IGNORE INTO mpce.keyword_free_association (keyword_1, keyword_2)
            SELECT k1.keyword_code AS keyword_1, k2.keyword_code AS keyword_2
            FROM manuscripts.keyword_free_associations AS ka
                LEFT JOIN manuscripts.keywords AS k1
                    ON k1.keyword = ka.keyword
                LEFT JOIN manuscripts.keywords AS k2
                    ON k2.keyword = ka.association
        """)
        cur.execute("""
            INSERT IGNORE INTO mpce.keyword_tree_association (keyword_1, keyword_2)
            SELECT k1.keyword_code AS keyword_1, k2.keyword_code AS keyword_2
            FROM manuscripts.keyword_tree_associations AS ka
                LEFT JOIN manuscripts.keywords AS k1
                    ON k1.keyword = ka.keyword
                LEFT JOIN manuscripts.keywords AS k2
                    ON k2.keyword = ka.association
        """)
        self.conn.commit()
        print(f'{cur.rowcount} keyword associations imported.')

        # Close cursor
        cur.close()

    def import_editions(self):
        """Imports edition data from manuscripts db"""

        # Open cursor
        cur = self.conn.cursor()

        print(f'Importing editions from `manuscripts.manuscript_books_editions`...')
        cur.execute("""
            INSERT INTO mpce.edition (
                edition_code, work_code, edition_status, edition_type,
                full_book_title, short_book_titles, translated_title,
                translated_language, languages, imprint_publishers,
                actual_publishers, imprint_publication_places,
                actual_publication_places, imprint_publication_years,
                actual_publication_years, pages, quick_pages,
                number_of_volumes, section, edition, book_sheets,
                notes, research_notes
            )
            SELECT book_code, super_book_code, edition_status,
                edition_type, full_book_title, short_book_titles,
                translated_title, translated_language, languages,
                stated_publishers, actual_publishers,
                stated_publication_places, actual_publication_places,
                stated_publication_years, actual_publication_years,
                pages, quick_pages, number_of_volumes, section,
                edition, book_sheets, notes, research_notes
            FROM manuscripts.manuscript_books_editions
        """)
        print(f'{cur.rowcount} editions imported into `mpce.edition`.')
        self.conn.commit()
        cur.close()

    def import_places(self):
        """Imports place data from manuscripts db"""

        cur = self.conn.cursor()

        cur.execute("""
            INSERT INTO mpce.place (
                place_code, name, alternative_names,
                town, C18_lower_territory, C18_sovereign_territory,
                C21_admin, C21_country, geographic_zone, BSR,
                HRE, EL, IFC, P, HE, HT, WT, PT, PrT,
                distance_from_neuchatel, latitude, longitude,
                geoname, notes
            )
            SELECT
                place_code, name, alternative_names,
                town, C18_lower_territory, C18_sovereign_territory,
                C21_admin, C21_country, geographic_zone, BSR,
                HRE, EL, IFC, P, HE, HT, WT, PT, PrT,
                distance_from_neuchatel, latitude, longitude,
                geoname, notes
            FROM manuscripts.places
        """)
        print(f'{cur.rowcount} places imported into `mpce.place`.')

        self.conn.commit()

    def import_stn(self):
        """Imports STN data from FBTEE-1"""

        cur = self.conn.cursor()

        # Port unchanged tables across
        print('Transferring unchanged STN data...')
        for mpce, man in self.UNCHANGED_TABLES.items():
            cur.execute(f'INSERT INTO {mpce} SELECT * FROM {man}')
            print(f'Data from `{man}` transferred to `{mpce}`.')
        self.conn.commit()

        print('Unchanged STN data imported. Importing transactions...')

        # Port transaction data across
        cur.execute("""
            CREATE TEMPORARY TABLE mpce.trans_type_key (
                name VARCHAR(255) PRIMARY KEY,
                id INT
            )
        """)
        cur.executemany("""
            INSERT INTO trans_type_key
            VALUES (%s, %s)
        """, seq_params=[(name, id) for name, id in self.TRANSACTION_CODING.items()])
        # Copy data across with new coding
        cur.execute("""
            INSERT INTO mpce.stn_transaction (
                transaction_code, order_code, page_or_folio_numbers,
                account_heading, direction, transaction_description, work_code,
                edition_code, stn_abbreviated_title, total_number_of_volumes,
                notes
            )
            SELECT
                t.transaction_code, t.order_code, t.page_or_folio_numbers,
                t.account_heading, tc.id, t.direction_of_transaction, t.super_book_code,
                t.book_code, t.stn_abbreviated_title, t.total_number_of_volumes,
                t.notes
            FROM manuscripts.transactions AS t
            LEFT JOIN mpce.trans_type_key AS tc
                ON t.direction_of_transaction LIKE tc.name
        """)
        self.conn.commit()
        print(f'{cur.rowcount} transactions ported into `mpce.stn_transaction` with new direction coding.') #pylint:disable=line-too-long;

        # Clients (need to parse dates)
        print(f'Importing client data, parsing dates ...')
        cur.execute('SELECT * FROM manuscripts.clients')
        clients = []
        for client in cur.fetchall():
            first = parse_date(client[9])
            last = parse_date(client[10])
            clients.append(client[:9] + (first,) + (last,) + client[-1:])
        cur.executemany("""
            INSERT INTO mpce.stn_client
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, clients)
        print(f'{cur.rowcount} clients inserted with parsed first and last dates.')
        self.conn.commit()

        # Finish
        cur.close()

    def import_new_tables(self):
        """Imports new MPCE data tables from manuscripts database."""

        cur = self.conn.cursor()

        print(f'Importing new datasets from manuscripts database ...')

        # L'estampillage de 1788
        cur.execute("""
            INSERT INTO mpce.stamping (
                ID, stamped_edition, permitted_dealer,
                attending_inspector, attending_adjoint,
                stamped_at_place, stamped_at_location_type,
                copies_stamped, volumes_stamped, date,
                ms_number, folio, citation, page_stamped,
                edition_notes, event_notes, article,
                date_entered, entered_by_user
            )
            SELECT
                ID, ID_EditionName, ID_DealerName,
                ID_AgentA, ID_AgentB,
                ID_PlaceName, EventLocation,
                IF(EventCopies = '', NULL, CONVERT(EventCopies, UNSIGNED INTEGER)),
                IF(EventVols = '', NULL, CONVERT(EventVols, UNSIGNED INTEGER)),
                EventDate,
                ID_Archive, EventFolioPage, EventCitation, EventPageStamped,
                EventNotes, EventOther, EventArticle,
                DateEntered, EventUser
            FROM manuscripts.manuscript_events
        """)
        self.conn.commit()
        print(f'{cur.rowcount} stampings copied into `mpce.stamping`.')

        # Illegal books
        # First clean up date and title fields
        cur.execute("""
            UPDATE manuscripts.manuscript_titles_illegal
            SET illegal_date = NULL
            WHERE illegal_date = 'No date available'
        """)

        # Import banned books
        cur.execute("""
            INSERT INTO mpce.banned_list_record (
                UUID, work_code, title, author, date, folio, notes
            )
            SELECT
                UUID, illegal_super_book_code, illegal_full_book_title, illegal_author_name,
                CONCAT(illegal_date, '-00-00'), illegal_folio, illegal_notes
            FROM manuscripts.manuscript_titles_illegal
            WHERE
                record_status <> 'DELETED' AND
                bastille_book_category = ''
        """)
        print(f'{cur.rowcount} banned books added to `mpce.banned_list_record`.')
        self.conn.commit()

        # Import bastille register records
        # Index to speed up import:
        cur.execute("""
            INSERT INTO mpce.bastille_register_record (
                UUID, work_code, title,
                author_name, imprint, publication_year,
                copies_found, current_volumes, total_volumes,
                category, notes
            )
            SELECT i.UUID, i.illegal_super_book_code, i.illegal_full_book_title,
                i.illegal_author_name, i.bastille_imprint_full, CONCAT(i.illegal_date, '-00-00'),
                i.bastille_copies_number, i.bastille_current_volumes, i.bastille_total_volumes,
                i.bastille_book_category, i.illegal_notes
            FROM manuscripts.manuscript_titles_illegal AS i
            WHERE CHAR_LENGTH(bastille_book_category) > 1
        """)
        print(f'{cur.rowcount} bastille register records added to `mpce.bastille_register_record`.')
        self.conn.commit()

        # Parisian stock auctions
        cur.execute("""
            INSERT INTO mpce.parisian_stock_auction (
                auction_id, ms_number, previous_owner, auction_reason, place
            )
            SELECT salesNumber, msNumber, Client_Code, code, Place_Code
            FROM manuscripts.manuscript_sales_events
        """)
        print(f'{cur.rowcount} stock auctions addded to `mpce.parisian_stock_auction`.')
        self.conn.commit()

        # Auction administrators
        auction_rgx = re.compile(r'(c[a-z][0-9]{3,4}) \((\w+)\)')

        # Get all the administrators
        cur.execute('SELECT salesNumber, ID_Agent FROM manuscripts.manuscript_sales_events')
        administrators = cur.fetchall()

        # Make dict of auction_roles
        cur.execute('SELECT * FROM auction_role')
        auction_roles = cur.fetchall()
        auction_roles = {role:id for id, role in auction_roles}

        # Split and flatten
        auction_administrator = []
        for sale, agents in administrators:
            # Extract data from string and append to list
            for agent in agents.split(','):
                mtch = auction_rgx.search(agent)
                if mtch:
                    administrator = mtch.group(1)
                    role = auction_roles[mtch.group(2)]
                else:
                    administrator = agent
                    role = None
                auction_administrator.append((sale, administrator, role))
        cur.executemany(
            """INSERT INTO mpce.auction_administrator
            VALUES (%s, %s, %s)""",
            seq_params=auction_administrator
        )
        print(f'{cur.rowcount} administration roles added to `mpce.auction_administrator`.')
        self.conn.commit()

        # Import individual sales
        cur.execute("""
            INSERT INTO mpce.parisian_stock_sale (
                ID, auction_id, purchaser,
                purchased_edition, sale_type,
                units_sold, units, volumes_traded,
                lot_price, date, folio,
                citation, article_number, edition_notes,
                event_notes, sale_notes
            )
            SELECT
                ss.ID, ss.ID_Sale_Agent, ss.ID_DealerName,
                ss.ID_EditionName, st.ID,
                ss.EventCopies,
                CASE WHEN EventCopiesType = 'packet' THEN 5
                    WHEN EventCopiesType = 'copies' THEN 3
                    WHEN EventCopiesType = 'privilege' THEN 8
                    WHEN EventCopiesType = 'plates' THEN 6
                    WHEN EventCopiesType = 'basket' THEN 4
                    WHEN EventCopiesType = 'vols' THEN 9
                    WHEN EventCopiesType = 'crate' THEN 2
                    ELSE NULL
                    END AS units,
                EventVols, EventLotPrice, EventDate, EventFolioPage,
                EventCitation, EventArticle, EventNotes,
                EventOther, EventMoreNotes
            FROM manuscripts.manuscript_events_sales AS ss
            LEFT JOIN mpce.sale_type AS st
                ON ss.EventType = st.type
        """)
        print(f'{cur.rowcount} sales added to `mpce.parisian_stock_sale`.')
        self.conn.commit()

        # Finish
        cur.close()

    def import_data_spreadsheets(self):
        """Imports major data spreadsheets from MPCE.

        NB: This function does not fully import the consignments data. The confiscation
        register signatories and the censors are left to self.resolve_agents()."""

        cur = self.conn.cursor()

        # Import consignments
        with path('mpcereform.spreadsheets', 'consignments.xlsx') as pth:
            print(f'Importing confiscations data from {pth} ...')
            consignments = load_workbook(pth, read_only=True, keep_vba=False)
        insert_params = []
        for row in consignments['Confiscations master'].iter_rows(min_row=2,
                                                                  max_col=45, values_only=True):
            insert_params.append({
                'ID': row[0],
                'UUID': str(uuid1()),
                'conf_reg_ms': row[1],
                'conf_reg_fol': row[2],
                'cust_reg_ms': row[3],
                'cust_reg_fol': row[4],
                '21935_fol': row[5],
                '21935_no': row[44],
                'date': row[6],
                'ship_no': row[7],
                'marque': row[8],
                'acquit': row[9],
                'stakeholder': row[32],
                'or_text': row[33],
                'or_code': row[34],
                'return_name': row[36],
                'return_agent': row[38],
                'return_town': row[40],
                'return_place': row[41],
                'notes': row[42]
            })

        cur.executemany("""
            INSERT INTO mpce.consignment (
                ID, UUID, confiscation_register_ms, confiscation_register_folio,
                customs_register_ms, customs_register_folio,
                ms_21935_folio, ms_21935_entry_no, shipping_number, marque,
                inspection_date, origin_text, origin_code,
                other_stakeholder, acquit_a_caution, returned_to_name,
                returned_to_agent, returned_to_town,
                returned_to_place, notes
            )
            VALUES (
                %(ID)s, %(UUID)s, %(conf_reg_ms)s, %(conf_reg_fol)s,
                %(cust_reg_ms)s, %(cust_reg_fol)s,
                %(21935_fol)s, %(21935_no), %(ship_no)s, %(marque)s,
                %(date)s, %(or_text)s, %(or_code)s,
                %(stakeholder)s, %(acquit)s, %(return_name)s,
                %(return_agent)s, %(return_town)s,
                %(return_place)s, %(notes)s
            )
        """, insert_params)
        print(f'{cur.rowcount} consignments imported into `mpce.consignment`.')
        self.conn.commit()

        # Import concerned agents for each consignment
        self._import_spreadsheet_agents(
            'mpce.consignment_addressee', consignments['Confiscations master'], cur, 10, 12)
        self._import_spreadsheet_agents(
            'mpce.consignment_signatory', consignments['Confiscations master'], cur, 27, 29)
        self._import_spreadsheet_agents(
            'mpce.consignment_handling_agent', consignments['Confiscations master'], cur, 17, 18)

        # Import permission simple
        with path('mpcereform.spreadsheets', 'permission_simple.xlsx') as pth:
            perm_simp = load_workbook(pth, read_only=True, keep_vba=False)
            print(f'Importing permission simple data from {pth} ...')

        perm_simp_grants = []
        for row in perm_simp['Licences'].iter_rows(min_row=2, max_row=1768, max_col=14, values_only=True):
            daw_wk, daw_ed, date, edn, _, _ = row[:6]
            licensee, _, _, _, l_cop, p_cop, spbk_conf, ed_conf = row[6:14]

            date = parse_date(date)

            perm_simp_grants.append(
                (daw_wk, daw_ed, date, edn, licensee,
                 l_cop, p_cop, spbk_conf, ed_conf)
            )

        cur.executemany("""
            INSERT INTO mpce.permission_simple_grant (
                dawson_work, dawson_edition, date_granted,
                edition_code, licensee, licensed_copies,
                printed_copies_estimate, work_confirmed,
                edition_confirmed
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, seq_params=perm_simp_grants)
        print(f'{cur.rowcount} licences imported into `mpce.permission_simple_grant`.')
        self.conn.commit()

        updated_edition_data = []
        for row in perm_simp['Editions'].iter_rows(min_row=2, max_row=1768, max_col=30, values_only=True):
            code, status, ed_type, _, full_title, short_title = row[:6]
            trans_title, trans_lang, lang, imprint_pub = row[6:10]
            act_pub, _, imp_place, act_place, _, stated_yrs = row[10:16]
            act_yrs, pge, quick_pg, vols, sec, edn, sheets = row[16:23]
            _, _, _, _, notes, research_notes, url = row[23:30]

            updated_edition_data.append((
                code, status, ed_type, full_title, short_title,
                trans_title, trans_lang, lang, imprint_pub,
                act_pub, imp_place, act_place, stated_yrs,
                act_yrs, pge, quick_pg, vols, sec, edn, sheets,
                notes, research_notes, url
            ))

        cur.executemany("""
            INSERT INTO mpce.edition (
                edition_code, edition_status, edition_type,
                full_book_title, short_book_titles,
                translated_title, translated_language,
                languages, imprint_publishers,
                actual_publishers, imprint_publication_places,
                actual_publication_places, imprint_publication_years,
                actual_publication_years, pages,
                quick_pages, number_of_volumes, section,
                edition, book_sheets, notes, research_notes,
                url
            )
            VALUES(
                %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                %s, %s, %s
            )
            ON DUPLICATE KEY UPDATE
                edition_status = VALUES(edition_status),
                edition_type = VALUES(edition_type),
                full_book_title = VALUES(full_book_title),
                short_book_titles = VALUES(short_book_titles),
                translated_title = VALUES(translated_title),
                translated_language = VALUES(translated_language),
                languages = VALUES(languages),
                imprint_publishers = VALUES(imprint_publishers),
                actual_publishers = VALUES(actual_publishers),
                imprint_publication_places = VALUES(imprint_publication_places),
                actual_publication_places = VALUES(actual_publication_places),
                pages = VALUES(pages),
                quick_pages = VALUES(quick_pages),
                number_of_volumes = VALUES(number_of_volumes),
                section = VALUES(section),
                edition = VALUES(edition),
                book_sheets = VALUES(book_sheets),
                notes = VALUES(notes),
                research_notes = VALUES(research_notes),
                url = VALUES(url)
        """, seq_params=updated_edition_data)
        print(f'{cur.rowcount} editions added or updated from permission simple spreadsheet.')
        self.conn.commit()

        # Import condemnations
        with path('mpcereform.spreadsheets', 'condemnations.xlsx') as pth:
            comdemn = load_workbook(pth, read_only=True, keep_vba=False)
            print(f'Importing condemnation data from {pth} ...')

        condemn_data = [row for row in comdemn['Sheet1'].iter_rows(
            min_row=2, max_row=114, max_col=6, values_only=True)]

        cur.executemany("""
            INSERT INTO mpce.condemnation (
                folio, title, notes, institution_text, date, other_judgment
            )
            VALUES (%s, %s, %s, %s, %s, %s)
        """, seq_params=condemn_data)
        print(f'{cur.rowcount} condemnations inserted into `mpce.condemnation`.')
        self.conn.commit()

        # Import Darnton sample
        with path('mpcereform.spreadsheets', 'CommandesLibrairesfrancais.xlsx') as pth:
            darnton_sample = load_workbook(pth, read_only=True, keep_vba=False)
            print(f'Importing additional STN order data from {pth} ...')

        darnton_data = []
        for row in darnton_sample['FicheSauvegarde'].iter_rows(min_row=2, max_row=3399, max_col=11, values_only=True):

            # Unpack row
            title, bk_format, volumes, author, num, date = row[:6]
            long_title, darn_id, ordered_by, _, notes = row[6:11]

            # Get client code
            if ordered_by in self.DARNTON_CLIENTS:
                ordered_by = self.DARNTON_CLIENTS[ordered_by]

            # Reformat date
            date = date.replace('?', '0')
            day = date[0:2]
            month = date[3:5]
            year = date[6:10]
            date = year + '-' + month + '-' + day

            darnton_data.append(
                (darn_id, title, bk_format, volumes, author, num, date,
                 long_title, ordered_by, notes)
            )

        cur.executemany("""
            INSERT INTO mpce.stn_darnton_sample_order (
                ID, title, format, volumes, author, num_ordered, date_ordered,
                edition_long_title, ordered_by, notes
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, seq_params=darnton_data)
        print(f'{cur.rowcount} book orders imported into `mpce.stn_darnton_sample_order`.')

        # Import provincial inspections
        with path('mpcereform.spreadsheets', 'provincial_inspections.xlsx') as pth:
            prov_insp = load_workbook(pth, read_only=True, keep_vba=False)
            print(f'Importing provincial inspections from {pth} ...')

        inspection_data = [row for row in prov_insp['Amalgamated sheet'].iter_rows(
            min_row=2, max_row=230, max_col=23, values_only=True)]

        cur.execute("""
            CREATE TEMPORARY TABLE mpce.prov_insp_temp (
                `ID` INT NOT NULL AUTO_INCREMENT,
                `ms_ref` VARCHAR(255),
                `folio` VARCHAR(255),
                `inspected_in` VARCHAR(255),
                `item` INT,
                `inspected_on` DATE,
                `ballot` VARCHAR(255),
                `consignment` VARCHAR(255),
                `acquit_a_caution` VARCHAR(255),
                `origin` VARCHAR(255),
                `author` TEXT,
                `title` TEXT,
                `imprint_place` TEXT,
                `imprint_publisher` TEXT,
                `imprint_date` CHAR(4),
                `volumes` VARCHAR(255),
                `format` VARCHAR(255),
                `languages` VARCHAR(255),
                `addressee` VARCHAR(255),
                `num_copies` INT,
                `inspected_by` TEXT,
                `decision` VARCHAR(255),
                `decision_date` DATE,
                `notes` TEXT,
                PRIMARY KEY(`ID`)
            )
        """)
        cur.executemany("""
            INSERT INTO prov_insp_temp (
                ms_ref, folio, inspected_in, item, inspected_on, ballot,
                consignment, acquit_a_caution, origin, author,
                title, imprint_place, imprint_publisher,
                imprint_date, volumes, format, languages, addressee,
                num_copies, inspected_by, decision, decision_date,
                notes
            )
            VALUES (
                %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                %s, %s, %s
            )
        """, seq_params=inspection_data)
        cur.execute("""
            INSERT INTO provincial_inspection (
                ID, ms_ref, folio, inspected_in, item,
                inspected_on, ballot, consignment, acquit_a_caution,
                origin, author, title, imprint_place,
                imprint_publisher, imprint_date, volumes,
                format, languages, addressee, num_copies,
                inspected_by, decision, decision_date, notes
            )
            SELECT
                tmp.ID, tmp.ms_ref, tmp.folio, insp_pl.place_code, tmp.item,
                tmp.inspected_on, tmp.ballot, tmp.consignment,
                tmp.acquit_a_caution, or_pl.place_code, tmp.author,
                tmp.title, tmp.imprint_place, tmp.imprint_publisher,
                tmp.imprint_date, tmp.volumes, tmp.format, tmp.languages,
                tmp.addressee, tmp.num_copies, tmp.inspected_by,
                tmp.decision, tmp.decision_date, tmp.notes
            FROM prov_insp_temp AS tmp
                LEFT JOIN mpce.place AS insp_pl
                    ON tmp.inspected_in = insp_pl.name
                LEFT JOIN mpce.place AS or_pl
                    ON tmp.origin = or_pl.name
        """)
        print(f'{cur.rowcount} events imported into `mpce.provincial_inspections`.')

        # Finish
        cur.close()

    def resolve_agents(self):
        """Resolves references to persons and corporate entities in the database.

        This method reforms all the agent data in the database,
        based on the information in the manuscripts database, and
        in the provided spreadsheets."""

        cur = self.conn.cursor()

        # Import basic agent data
        # Lenghten all person_codes by two digits.
        print('Importing existing agent data...')
        cur.execute("""
            SELECT
                CONCAT('id00', RIGHT(person_code, 4)), person_name, sex, title,
                other_names, designation, status, birth_date, death_date, notes
            FROM manuscripts.people
        """)
        people = []
        for person in cur.fetchall():
            birth = parse_date(person[7])
            death = parse_date(person[8])
            people.append(person[:7] + (birth,) + (death,) + person[-1:])
        cur.executemany("""
            INSERT INTO mpce.agent (
                agent_code, name, sex, title, other_names,
                designation, status, start_date, end_date,
                notes
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, people)
        print(f'{cur.rowcount} agents imported from `manuscripts.people` into `mpce.agent`.')

        # Get client-agent data from STN database
        print('Importing stn client-agent relationships...')
        cur.execute("""
            INSERT INTO mpce.stn_client_agent
            SELECT client_code, CONCAT('id00', RIGHT(person_code, 4))
            FROM manuscripts.clients_people
        """)
        print(f'{cur.rowcount} relationships inserted into `mpce.stn_client_agent`.')

        # Import agent metadata
        cur.execute("""
            INSERT INTO mpce.profession
            SELECT *
            FROM manuscripts.professions
        """)
        cur.execute("""
            INSERT INTO mpce.agent_profession
            SELECT CONCAT('id00', RIGHT(person_code, 4)), profession_code
            FROM manuscripts.people_professions
        """)
        print((
            'Professions and assignments imported from `manuscripts.professions` '
            'and `manuscripts.people_professions`.'
        ))

        # The permission simple sheet contains some new professions
        with path('mpcereform.spreadsheets', 'permission_simple.xlsx') as pth:
            print(f'Importing new profession data from {pth}')
            p_simple = load_workbook(pth, read_only=True, keep_vba=False)
        new_professions = [row for row in p_simple['New Professions'].iter_rows(
            min_row=2, values_only=True) if row[0] is not None]
        cur.executemany("""
            INSERT IGNORE INTO mpce.profession (
                profession_type, profession_code, profession_group, economic_sector
            )
            VALUES (%s, %s, %s, %s)
        """, new_professions)
        print(f'{cur.rowcount} new professions imported.')
        self.conn.commit()

        # Now all agent_codes (person_codes) have been imported, as have profession codes.

        # RESOLVE AUTHORS:

        # Import author data
        print('Resolving authors...')
        with path('mpcereform.spreadsheets', 'author_person.xlsx') as pth:
            author_person = load_workbook(pth, read_only=True, keep_vba=False)
            print(f'Author-agent assignments loaded from {pth}')
        # Get list of all authors who already have agent codes
        assigned_authors = []
        for row in author_person['author_person'].iter_rows(min_row=2, values_only=True):
            # If the match is correct...
            if row[7] == 'Y':
                # ... append (agent_code, author_code)
                # Assumes that the workbook has the following columns in sheet 0:
                # agent_code, client_code, agent_name, author_code, author_name, osa, cosine, correct, notes
                assigned_authors.append((row[0], row[3]))
        # Create temporary author_agent table
        cur.execute("""
            CREATE TEMPORARY TABLE mpce.author_agent (
                agent_code VARCHAR(255),
                author_code VARCHAR(255),
                PRIMARY KEY(author_code, agent_code)
            )
        """)
        cur.executemany("""
            INSERT INTO mpce.author_agent
            VALUES (%s, %s)
        """, seq_params=assigned_authors)
        self.conn.commit()
        print(f'{cur.rowcount} authors with agent_codes found in spreadsheet.')

        # Create new agents for all authors without an agent_code
        cur.execute("""
            SELECT ma.author_name, ma.author_code
            FROM manuscripts.manuscript_authors AS ma
            LEFT JOIN mpce.author_agent AS aa
                ON aa.author_code = ma.author_code
            WHERE aa.agent_code IS NULL
        """)
        # There are duplicate authors...
        # Create dict of authors
        unassigned_auths = {code:name for name, code in cur.fetchall()}
        # Get unique names, and assign agent_codes
        unique_names = set(unassigned_auths.values())
        num = len(unique_names)
        new_agent_codes = self._get_code_sequence('mpce.agent', 'agent_code', num, cur)
        name_code = {name: code for name, code in zip(
            unique_names, new_agent_codes)}
        cur.executemany("""
            INSERT INTO mpce.agent (agent_code, name)
            VALUES (%s, %s)
        """, seq_params=[(code, name) for name, code in name_code.items()])
        # Now map these new agent codes back onto author table
        auth_agent = [(author_code, name_code[name]) for author_code, name in unassigned_auths.items()]
        cur.executemany("""
            INSERT INTO mpce.author_agent (author_code, agent_code)
            VALUES (%s, %s)
        """, seq_params=auth_agent)
        print(f'{cur.rowcount} authors assigned new agent_codes...')
        self.conn.commit()
        # Now import authorship data
        cur.execute("""
            INSERT INTO mpce.edition_author (
                edition_code, author, author_type, certain
            )
            SELECT ba.book_code, aa.agent_code, at.id, ba.certain
                FROM manuscripts.manuscript_books_authors AS ba
                LEFT JOIN mpce.author_agent AS aa
                    ON ba.author_code = aa.author_code
                LEFT JOIN mpce.author_type AS at
                    ON ba.author_type LIKE at.type
        """)
        print(f'All authors resolved into agents. {cur.rowcount} authorship attributions imported into `mpce.edition_author`.')
        self.conn.commit()

        # Apply new profession code to all authors
        cur.execute("""
            INSERT IGNORE INTO mpce.agent_profession
            SELECT
                ea.author,
                CASE WHEN ea.author_type = 1 THEN 'pf014'
                    WHEN ea.author_type = 2 THEN 'pf014'
                    WHEN ea.author_type = 3 THEN 'pf310'
                    WHEN ea.author_type = 4 THEN 'pf227'
                    ELSE NULL
                END
            FROM mpce.edition_author AS ea
        """)
        print(
            f'{cur.rowcount} profession codes assigned to "aucteurs", "redacteurs" and "traducteurs".')
        self.conn.commit()

        # RESOLVE CLIENTS

        # Create a combined list of all clients
        print('Finding client codes...')
        cur.execute("""
            CREATE TEMPORARY TABLE mpce.all_clients (
                client_code VARCHAR(255) PRIMARY KEY,
                name VARCHAR(255),
                alt_name VARCHAR(255),
                prof_codes VARCHAR(255),
                place_codes VARCHAR(255),
                gender VARCHAR(255),
                title VARCHAR(255),
                notes TEXT,
                corporate BIT(1)
            )
        """)
        print('Scanning STN clients ...')
        cur.execute("""
            INSERT INTO mpce.all_clients (
                client_code, name, gender, corporate, notes
            )
            SELECT client_code, client_name, gender, partnership, notes
            FROM mpce.stn_client
        """)
        print('Scanning `manuscripts.manuscript_dealers`...')
        cur.execute("""
            INSERT INTO mpce.all_clients (
                client_code, name, alt_name, prof_codes, place_codes, notes
            )
            SELECT
                Client_Code, Dealer_Name, Alternative_Name, Profession_Code,
                Place_Code, Notes
            FROM manuscripts.manuscript_dealers
            ON DUPLICATE KEY UPDATE mpce.all_clients.notes = CONCAT(IFNULL(mpce.all_clients.notes, ''), ' Stock Sales Notes: ', manuscripts.manuscript_dealers.notes)
        """)
        print('Scanning `manuscripts.manuscript_agents_inspectors`...')
        cur.execute("""
            INSERT INTO mpce.all_clients (
                client_code, name, place_codes, notes
            )
            SELECT
                Client_Code, Agent_Name, Place_Code, Notes
            FROM manuscripts.manuscript_agents_inspectors
            ON DUPLICATE KEY UPDATE mpce.all_clients.notes = CONCAT(IFNULL(mpce.all_clients.notes, ''), ' Estampillage Notes: ', manuscripts.manuscript_agents_inspectors.notes)
        """)

        # New clients in consignments workbook
        with path('mpcereform.spreadsheets', 'consignments.xlsx') as pth:
            print(f'Scanning {pth} ...')
            consignments = load_workbook(pth, read_only=True, keep_vba=False)
        consignment_clients = {}
        for row in consignments['People final'].iter_rows(min_row=2, values_only=True):
            if isinstance(row[3], str) or ~isinstance(row[2], str):
                continue

            # Split codes and names for multi-person cells
            codes = row[3].split(';')
            names = row[2].split(';')

            # Get other data (these columns are never split)
            notes = row[4]
            title = row[5]
            place = row[7]

            for code, name in zip(codes, names): # zip() deals with different-length sequences
                code = code.strip()
                if code not in consignment_clients:
                    consignment_clients[code] = (code, name, notes, title, place)
        cur.executemany("""
            INSERT INTO mpce.all_clients (
                client_code, name, notes, title, place_codes
            )
            VALUES (%s, %s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE mpce.all_clients.notes = CONCAT(IFNULL(mpce.all_clients.notes, ''), ' Confiscations notes: ', VALUES(notes))
        """, seq_params=consignment_clients.values())

        # New clients in permission simple
        with path('mpcereform.spreadsheets', 'permission_simple.xlsx') as pth:
            print(f'Scanning {pth} ...')
            per_simp = load_workbook(pth, read_only=True, keep_vba=False)
        cur.executemany("""
            INSERT INTO mpce.all_clients (
                client_code, name, alt_name, gender, prof_codes, place_codes, notes
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE mpce.all_clients.notes = CONCAT(IFNULL(mpce.all_clients.notes, ''), ' Permission simple notes: ', VALUES(notes))
        """, [(r[0], r[1], r[2], r[3], r[5], r[7], r[8])
              for r in per_simp['Clients'].iter_rows(min_row=2, max_row=249, values_only=True)])
        self.conn.commit()
        cur.execute('SELECT COUNT(client_code) FROM mpce.all_clients')
        print(f'{cur.fetchone()[0]} clients found across all datasets.')

        # Insert data on new agents

        # Create temporary client_agent table from stn_client_agent
        # Remember to exclude partnerships--these relationships will go in the
        # 'is_member_of' table
        print('Converting clients to agents ...')
        cur.execute("""
            CREATE TEMPORARY TABLE mpce.client_agent (
                `client_code` CHAR(6) NOT NULL,
                `agent_code` CHAR(8) NOT NULL,
                PRIMARY KEY (`client_code`, `agent_code`)
            )
        """)

        # Get existing client-person data
        print('Scanning mpce.stn_client_agent ...')
        cur.execute("""
            INSERT INTO mpce.client_agent
            SELECT sca.client_code, sca.agent_code
            FROM mpce.stn_client_agent AS sca
                LEFT JOIN mpce.stn_client AS sc
                    ON sca.client_code = sc.client_code
            WHERE sc.partnership IS FALSE
        """)

        # Import new agents from `clients_without_person_codes.xlsx`
        with path('mpcereform.spreadsheets', 'clients_without_person_codes.xlsx') as pth:
            print(f'Creating new agents according from data in {pth} ...')
            new_stn_clients = load_workbook(
                pth, read_only=True, keep_vba=False)

        new_cl_ls = []
        for row in new_stn_clients['clients_without_person_codes'].iter_rows(min_row=2, values_only=True):
            client_code = row[0]
            client_name = row[1]
            if row[2] == 'Y':
                corporate = False
            elif row[3] == 'Y':
                corporate = True
            else:
                continue
            notes = row[4]
            new_cl_ls.append((client_code, client_name, corporate, notes))
        new_cl_agts = self._get_code_sequence(
            'mpce.agent', 'agent_code', len(new_cl_ls), cur)

        cur.executemany("""
            INSERT INTO mpce.agent (agent_code, name, corporate_entity, notes)
            VALUES (%s, %s, %s, %s)
        """, seq_params=[(code, name, corp, notes) for code, (client, name, corp, notes) in zip(new_cl_agts, new_cl_ls)])
        print(f'{cur.rowcount} new agents created.')
        cur.executemany("""
            INSERT INTO mpce.stn_client_agent (client_code, agent_code)
            VALUES (%s, %s)
        """, seq_params=[(client, code) for code, (client, name, corp, notes) in zip(new_cl_agts, new_cl_ls)])
        print(f'{cur.rowcount} new relationships inserted into `stn_client_agent`')
        cur.executemany("""
            INSERT INTO mpce.client_agent (client_code, agent_code)
            VALUES (%s, %s)
        """, seq_params=[(client, code) for code, (client, name, corp, notes) in zip(new_cl_agts, new_cl_ls)])
        self.conn.commit()

        # Generate new agent codes
        # The XOR allows the creation of new agent_codes for clients
        # that are linked to an agent of a different type to themselves.
        cur.execute("""
            SELECT
                ac.client_code, ac.name, ac.alt_name,
                ac.prof_codes, ac.place_codes, ac.gender,
                ac.notes, ac.corporate, ac.title
            FROM mpce.all_clients AS ac
                LEFT JOIN mpce.client_agent AS ca
                    ON ac.client_code = ca.client_code
                LEFT JOIN mpce.agent AS a
                    ON ca.agent_code = a.agent_code
            WHERE
                (
                    ca.agent_code IS NULL OR
                    (ac.corporate XOR a.corporate_entity)
                )
                AND ac.name NOT LIKE 'null'
        """)
        new_agents = cur.fetchall()
        num_new_codes = len(new_agents)
        print(f'Assigning new agent codes to {num_new_codes} clients ...')
        code_list = self._get_code_sequence(
            'mpce.agent', 'agent_code', num_new_codes, cur)
        cur.executemany("""
            INSERT INTO mpce.client_agent (client_code, agent_code)
            VALUES (%s, %s)
        """, seq_params=[(client[0], code) for client, code in zip(new_agents, code_list)])

        # Process new_agent data:
        processed_agents = []
        new_prof_assigns = []
        new_place_assigns = []
        for client, agent_code in zip(new_agents, code_list):
            _, name, alt_names, prof_codes = client[0:4]
            place_codes, gender, notes, corporate, title = client[4:9]

            # Process gender
            if gender is not None:
                if gender.lower().startswith('mixed'):
                    gender = None
                    corporate = True
                elif gender.lower().startswith('m'):
                    gender = 'M'
                elif gender.lower().startswith('f'):
                    gender = 'F'

            # Process addresses
            if place_codes is not None:
                for place in re.split(r',\s*|;\s*', place_codes):
                    new_place_assigns.append((agent_code, place))

            # Process professions
            if prof_codes is not None:
                for prof in re.split(r',\s*|;\s*', prof_codes):
                    new_prof_assigns.append((agent_code, prof))

            # Add processed agent to list
            processed_agents.append(
                (agent_code, name, alt_names, gender, corporate, title)
            )

        # Generate new agents
        cur.executemany("""
            INSERT INTO mpce.agent (
                agent_code, name, other_names, sex, corporate_entity, title
            )
            VALUES (%s, %s, %s, %s, %s, %s)
        """, seq_params=processed_agents)
        print(f'{cur.rowcount} new agents added to `mpce.agent`.')
        self.conn.commit()

        # Assign places to new agents:
        # Using stn address data
        cur.execute("""
            INSERT INTO mpce.agent_address (agent_code, place_code, address)
            SELECT ca.agent_code, addr.place_code, addr.address
            FROM manuscripts.clients_addresses AS addr
                LEFT JOIN mpce.client_agent AS ca
                    ON addr.client_code = ca.client_code
        """)
        print(f'{cur.rowcount} addresses imported from `manuscripts.clients_addresses`.')

        # Using generated python list
        # I tried to do this by creating a temporary index on agent_code and place_code,
        # and then doing an INSERT IGNORE, but for some reason mysql.connector kept
        # throwing an IntegrityError. So here's a hacky version...
        cur.execute('SELECT agent_code, place_code FROM mpce.agent_address')
        existing_addresses = set(cur.fetchall())
        new_place_assigns = [assign for assign in new_place_assigns if assign not in existing_addresses]
        cur.executemany("""
            INSERT INTO mpce.agent_address (agent_code, place_code)
            VALUES (%s, %s)
        """, seq_params=new_place_assigns)
        print(f'{cur.rowcount} addresses imported from new datasets.')
        self.conn.commit()

        # Assign professions to new agents:
        cur.executemany("""
            INSERT IGNORE INTO mpce.agent_profession
            VALUES (%s, %s)
        """, seq_params=new_prof_assigns)
        print(f'{cur.rowcount} new professions assigned to agents')
        self.conn.commit()

        # Update notes:
        cur.execute("""
            UPDATE mpce.agent AS a, mpce.all_clients AS ac, mpce.client_agent AS ca
            SET a.notes = TRIM(CONCAT(IFNULL(a.notes, ''), ' ', ac.notes))
            WHERE a.agent_code = ca.agent_code AND ca.client_code = ac.client_code
        """)
        print(f'Notes concatenated from different datasets for {cur.rowcount} agents.')
        self.conn.commit()

        # Use temporary join table to replace client codes throughout db:
        cur.execute("""
            UPDATE mpce.consignment AS tbl
            LEFT JOIN mpce.client_agent AS ca ON tbl.other_stakeholder = ca.client_code
            SET tbl.other_stakeholder = ca.agent_code
        """)
        print(
            f'{cur.rowcount} other_stakeholders in `mpce.consignment` resolved into agent_codes.')
        cur.execute("""
            UPDATE mpce.consignment AS tbl
            LEFT JOIN mpce.client_agent AS ca ON tbl.returned_to_agent = ca.client_code
            SET tbl.returned_to_agent = ca.agent_code
        """)
        print(f'{cur.rowcount} returned_to_agents in `mpce.consignment` resolved into agent_codes.')

        # Store all_collectors and all_censors as strings
        cur.execute("""
            CREATE TEMPORARY TABLE all_collectors (
                `consignment` INT,
                `agent_code` CHAR(8),
                `text` VARCHAR(255),
                PRIMARY KEY (`consignment`,`agent_code`)
            );
        """)
        cur.execute("""
            CREATE TEMPORARY TABLE all_censors (
                `consignment` INT,
                `agent_code` CHAR(8),
                `text` VARCHAR(255),
                PRIMARY KEY (`consignment`,`agent_code`)
            );
        """)

        # Get collector and censor data from consignments workbook
        self._import_spreadsheet_agents(
            'all_collectors', consignments['Confiscations master'], cur, 24, 25)
        self._import_spreadsheet_agents(
            'all_censors', consignments['Confiscations master'], cur, 20, 21)
        # Splice into consignment table
        cur.execute("""
            UPDATE mpce.consignment AS cons
            LEFT JOIN (
                SELECT
                    consignment,
                    GROUP_CONCAT(
                        CONCAT(all_c.text, ' (', ac.agent_code, ')')
                        SEPARATOR '; '
                    ) AS out_string
                FROM all_collectors AS all_c
                    LEFT JOIN client_agent AS ac
                        ON all_c.agent_code = ac.client_code
                GROUP BY consignment
            ) AS colls
            ON colls.consignment = cons.ID
            SET cons.all_collectors = colls.out_string
        """)
        cur.execute("""
            UPDATE mpce.consignment AS cons
            LEFT JOIN (
                SELECT
                    consignment,
                    GROUP_CONCAT(
                        CONCAT(all_c.text, ' (', ac.agent_code, ')')
                        SEPARATOR '; '
                    ) AS out_string
                FROM all_censors AS all_c
                    LEFT JOIN client_agent AS ac
                        ON all_c.agent_code = ac.client_code
                GROUP BY consignment
            ) AS colls
            ON colls.consignment = cons.ID
            SET cons.all_censors = colls.out_string
        """)
        print(f'Censor and collector data imported into `mpce.consignment`.')

        cur.execute("""
            UPDATE mpce.consignment_addressee AS tbl
            LEFT JOIN mpce.client_agent AS ca ON tbl.agent_code = ca.client_code
            SET tbl.agent_code = ca.agent_code
        """)
        print(f'{cur.rowcount} client codes in `mpce.consignment_addressee` resolved into agent_codes.')
        cur.execute("""
            UPDATE mpce.consignment_signatory AS tbl
            LEFT JOIN mpce.client_agent AS ca ON tbl.agent_code = ca.client_code
            SET tbl.agent_code = ca.agent_code
        """)
        print(f'{cur.rowcount} client codes in `mpce.consignment_signatory` resolved into agent_codes.')
        cur.execute("""
            UPDATE mpce.consignment_handling_agent AS tbl
            LEFT JOIN mpce.client_agent AS ca ON tbl.agent_code = ca.client_code
            SET tbl.agent_code = ca.agent_code
        """)
        print(f'{cur.rowcount} client codes in `mpce.consignment_handling_agent` resolved into agent_codes.')

        cur.execute("""
            UPDATE mpce.stamping AS tbl
            LEFT JOIN mpce.client_agent AS ca ON tbl.permitted_dealer = ca.client_code
            SET tbl.permitted_dealer = ca.agent_code
        """)
        cur.execute("""
            UPDATE mpce.stamping AS tbl
            LEFT JOIN mpce.client_agent AS ca ON tbl.attending_inspector = ca.client_code
            SET tbl.attending_inspector = ca.agent_code
        """)
        cur.execute("""
            UPDATE mpce.stamping AS tbl
            LEFT JOIN mpce.client_agent AS ca ON tbl.attending_adjoint = ca.client_code
            SET tbl.attending_adjoint = ca.agent_code
        """)
        print(f'{cur.rowcount} client codes in `mpce.stamping` resolved into agent_codes.')

        cur.execute("""
            UPDATE mpce.parisian_stock_auction AS tbl
            LEFT JOIN mpce.client_agent AS ca ON tbl.previous_owner = ca.client_code
            SET tbl.previous_owner = ca.agent_code
        """)
        print(f'{cur.rowcount} client codes in `mpce.parisian_stock_auction` resolved into agent_codes.')
        cur.execute("""
            UPDATE mpce.auction_administrator AS tbl
            LEFT JOIN mpce.client_agent AS ca ON tbl.administrator_id = ca.client_code
            SET tbl.administrator_id = ca.agent_code
        """)
        print(f'{cur.rowcount} client codes in `mpce.auction_administrator` resolved into agent_codes.')
        cur.execute("""
            UPDATE mpce.parisian_stock_sale AS tbl
            LEFT JOIN mpce.client_agent AS ca ON tbl.purchaser = ca.client_code
            SET tbl.purchaser = ca.agent_code
        """)
        print(f'{cur.rowcount} client codes in `mpce.parisian_stock_sale` resolved into agent_codes.')
        self.conn.commit()
        cur.execute("""
            UPDATE mpce.permission_simple_grant AS tbl
            LEFT JOIN mpce.client_agent AS ca ON tbl.licensee = ca.client_code
            SET tbl.licensee = ca.agent_code
        """)
        print(f'{cur.rowcount} client codes in `mpce.permission_simple_grant` resolved into agent_codes.')
        self.conn.commit()

        # Populate 'is member of' from stn data
        cur.execute("""
            INSERT INTO mpce.is_member_of (member, corporate_entity)
            SELECT sca.agent_code, ca.agent_code
            FROM mpce.stn_client_agent AS sca
            LEFT JOIN mpce.client_agent AS ca
                ON sca.client_code = ca.client_code
            LEFT JOIN mpce.stn_client AS cl_orig
                ON sca.client_code = cl_orig.client_code
            LEFT JOIN mpce.agent AS a
                ON sca.agent_code = a.agent_code
            WHERE
                cl_orig.partnership IS TRUE AND
                a.corporate_entity IS NOT TRUE
        """)
        print(f'{cur.rowcount} memberships of corporate entities imported from STN data.')
        self.conn.commit()

        # Finish
        cur.close()

    def build_indexes(self):
        """Builds key indexes for common queries."""

        # QUERY: add foreign key constraints at this stage?

        pass #pylint:disable=unnecessary-pass;

    def summarise(self):
        """Outputs summary statistics about the database."""

        cur = self.conn.cursor()

        print('\nMPCE data import complete.\n')
        print('SUMMARY STATISTICS:\n========================\n')

        # Works
        cur.execute('SELECT COUNT(work_code) FROM mpce.work')
        print(f'Distinct works: {cur.fetchone()[0]}, which have been assigned')
        cur.execute('SELECT COUNT(*) FROM mpce.work_keyword')
        print(f'     {cur.fetchone()[0]} keywords from a set of')
        cur.execute('SELECT COUNT(keyword_code) FROM mpce.keyword')
        print(f'     {cur.fetchone()[0]} categories devised by the project')

        print('')

        # Editions
        cur.execute('SELECT COUNT(edition_code) FROM mpce.edition')
        print(f'Distinct editions: {cur.fetchone()[0]}, produced by')
        cur.execute("""
            SELECT at.type, COUNT(*)
            FROM mpce.edition_author AS ea
                LEFT JOIN mpce.author_type AS at
                    ON ea.author_type = at.ID
            GROUP BY ea.author_type
        """)
        for auth_type, num in cur.fetchall():
            if auth_type is None:
                continue
            if auth_type in {'Primary', 'Secondary'}:
                print(f'     {num} {auth_type} authors')
            else:
                print(f'     {num} {auth_type}s')

        print('')

        # Agents:
        cur.execute("SELECT COUNT(agent_code), SUM(corporate_entity) FROM mpce.agent")
        agents, entities = cur.fetchone()
        print(f'Distinct agents: {agents}, of which')
        print(f'     {agents - entities} are persons')
        print(f'     {entities} are corporate entities')
        cur.execute('SELECT COUNT(*) FROM mpce.stn_client_agent')
        print(f'     {cur.fetchone()[0]} were clients of the STN')

        print('')

        # Places:
        cur.execute("SELECT COUNT(place_code) FROM mpce.place")
        print(f'Distinct places: {cur.fetchone()[0]}')

        print('')

        # Events
        events = {}
        cur.execute("SELECT COUNT(*) FROM mpce.banned_list_record")
        events['banned by the authorities'] = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM mpce.bastille_register_record")
        events['sequestered in the Bastille'] = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM mpce.condemnation")
        events['condemned by the authorities'] = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM mpce.consignment")
        events['intercepted in a suspect consignment by Paris customs'] = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM mpce.parisian_stock_sale")
        events['sold, or the right to print them transferred, at the Paris stock sales'] = cur.fetchone()[
            0]
        cur.execute("SELECT COUNT(*) FROM mpce.permission_simple_grant")
        events['licensed under the permission simple'] = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM mpce.provincial_inspection")
        events['inspected by provincial authorities'] = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM mpce.stamping")
        events['stamped to legalise their sale, though they were pirated'] = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM mpce.stn_transaction")
        events[(
            'bought, sold, sent, returned, printed, warehoused or otherwise\n'
            '         dealt with by the Société Typographique de Neuchâtel'
        )] = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM mpce.stn_darnton_sample_order")
        events['ordered by one of Robert Darnton\'s selected buyers'] = cur.fetchone()[0]

        print(f'All of which were involved in\n')
        print(f'     {sum(events.values())} distinct events\n')
        print(f'in the history of the book, including...\n')

        for key, val in events.items():
            print(f'     {val} times books were {key}')

        cur.close()

    # Utility methods
    def _get_code_sequence(self, table, column, num, cursor=None):
        """Return a list of the next n free codes.

        Arguments:
        ==========
            table (str): name of table to be queried
            column (str): name of column to be queried
            num (int): number of new ids to be generated
            cursor (MySQLCursor): a cursor, if you don't wish to create a new one

        Returns:
        ==========
            A sequence of n new codes
        """

        # Regexes
        num_extr_rgx = re.compile(r'[1-9]\d*') # Extract numerica part of id
        frame_rgx = re.compile(r'[a-z]+(?=0)') # To find frame

        # Get a cursor
        if cursor is not None:
            cur = cursor
        else:
            cur = self.conn.cursor()

        # Retrieve agent codes, strip 'id' and leading 0s, convert to int
        cur.execute(f'SELECT {column} FROM {table}')
        codes = cur.fetchall()

        # Work out the frame:
        alpha = frame_rgx.match(codes[0][0]).group(0)
        n_digits = len(codes[0][0]) - len(alpha)
        frame = ''.join([alpha] + ['0' for n in range(n_digits)])

        # Extract numeric part of codes
        codes = [int(num_extr_rgx.search(id).group(0))
                 for (id,) in codes]

        if cursor is None:
            cur.close()

        # Get the maximum numeric id
        next_id = max(codes) + 1

        # Return list of codes
        return [frame[:-len(str(id))] + str(id) for id in range(next_id, next_id + num)]

    def _import_spreadsheet_agents(self, table, worksheet, cursor, text_col, code_col):
        """Custom method for consignments workbook."""
        agents = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            agent_id = row[0]
            names, codes = row[text_col], row[code_col]
            if isinstance(names, str) or ~isinstance(codes, str):
                continue
            else:
                names = names.split(';')
                codes = codes.split(';')
            for name, code in zip(names, codes):
                agents.append((agent_id, code.strip(), name.strip()))
        cursor.executemany((
            f'INSERT INTO {table} (consignment, agent_code, text) '
            'VALUES (%s, %s, %s)'
        ), agents)
        print(f'{cursor.rowcount} agency relations inserted into `{table}`.')
        self.conn.commit()

    def _get_auto_increment(self, table, column, cur=None):
        """Returns the frame and next id value for the nominated id column."""
        if cur is None:
            cur = self.conn.cursor()

        # Regexes
        num_extr_rgx = re.compile(r'[1-9]\d*') # Extract numerical part of id
        frame_rgx = re.compile(r'[a-z]+(?=0)') # To find frame

        cur.execute(f'SELECT {column} FROM {table}')
        ids = cur.fetchall()

        # The next id is the max value + 1
        # Remember each record is a tuple
        next_id = max([int(num_extr_rgx.search(record[0]).group(0)) for record in ids]) + 1

        # Get the frame from any of the values
        # For legacy reasons, some of these id columns have different frames
        # This function looks at the first value, so gets the original frame
        frame = frame_rgx.search(ids[0][0]).group(0)

        return next_id, frame
